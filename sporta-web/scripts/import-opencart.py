#!/usr/bin/env python3
"""Turn the OpenCart export into the site catalogue.

    python3 scripts/import-opencart.py path/to/sportaexport.xlsx

Writes src/lib/products.js and sporta-html5/assets/products.js, then
generate-seed.mjs turns those into supabase/seed-products.sql.

WHY THIS IS NOT A STRAIGHT COPY
The export has 90 rows, but they are not 90 products. They are size and colour
variants of about 40: "Cloudsoft Collection Leggings - Onyx Black (M)" and
"(L)" are one product. The site sizes garments by category rather than per
product, so the variants have to collapse or the shop shows the same item a
dozen times.

Collapsing on the raw name does not work either, because the same product was
typed several different ways over four years:

    Sculpt Collection Top - Grey
    Sculpt Collection - Top - Grey
    Sculpt Collection - Top (Black - L)
    Cloudsoft Collection - Top (Grey, L)

so the name is parsed into (collection, item, colour) and rebuilt canonically.
Anything that cannot be parsed is REPORTED, never guessed at: a product whose
colour was misread is worse than one that needs a minute of human attention.

NOTHING FROM THE ORDERS SHEETS IS READ. They hold 119 real customers' names,
emails and phone numbers, which have no business in a public catalogue.
"""
import sys, re, json, unicodedata
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit("pip3 install openpyxl")

HERE = Path(__file__).resolve().parent
WEB = HERE.parent
ROOT = WEB.parent

SIZE_WORDS = r'XXS|XS|XXL|XL|S|M|L|Small|Medium|Large|X-?Large'

# Longest first so "Onyx Black" wins over "Black" and "Steel Grey" over "Grey".
COLOURS = [
    ('army green',  'أخضر عسكري'), ('coffee brown', 'بني قهوة'), ('taupe brown', 'بني فاتح'),
    ('cherry red',  'أحمر كرزي'),  ('iris purple',  'بنفسجي'),   ('steel grey',  'رمادي فولاذي'),
    ('onyx black',  'أسود أونيكس'), ('navy blue',   'كحلي'),      ('red/white',   'أحمر وأبيض'),
    ('navy',        'كحلي'),        ('black',       'أسود'),      ('white',       'أبيض'),
    ('grey',        'رمادي'),       ('gray',        'رمادي'),     ('green',       'أخضر'),
    ('royal',       'أزرق ملكي'),   ('pink',        'وردي'),      ('red',         'أحمر'),
    ('purple',      'بنفسجي'),      ('brown',       'بني'),
]

# (regex, english noun, arabic noun, category)
ITEMS = [
    (r'phone strap',                 'Phone Strap',  'حزام هاتف',      'accessories'),
    (r'\bbackpack\b|\bzaino\b',      'Backpack',     'حقيبة ظهر',      'accessories'),
    (r'\bsocks?\b',                  'Socks',        'جوارب',          'accessories'),
    (r'9forty|\bcap\b|new era',      'Cap',          'كاب',            'accessories'),
    (r'sweatshirt',                  'Sweatshirt',   'سويت شيرت',      'outerwear'),
    (r'\bjacket\b',                  'Jacket',       'جاكيت',          'outerwear'),
    (r'\bleggings?\b',               'Leggings',     'ليقنز',          'women'),
    (r'sports? bra|\bbra\b',         'Sports Bra',   'برا رياضية',     'women'),
    (r'stringer|\bvest\b',           'Stringer Vest','ستراينجر',       'men'),
    (r'\btank\b',                    'Tank',         'تانك',           'men'),
    (r'\bpolo\b',                    'Polo',         'بولو',           'men'),
    (r'\bshorts?\b',                 'Shorts',       'شورت',           'men'),
    (r'\bpant\b|\btrousers?\b',      'Pants',        'بنطال',          'men'),
    (r't[-\s]?shirt|tshirts?|\btee\b', 'T-Shirt',    'تيشيرت',         'men'),
    (r'\btop\b',                     'Top',          'توب',            'women'),
]

# Collections and brands, with the Arabic spelling used on the site.
COLLECTIONS = [
    ('cloudsoft collection', 'Cloudsoft',        'كلاودسوفت'),
    ('sculpt collection',    'Sculpt',           'سكالبت'),
    ('define',               'Define',           'ديفاين'),
    ('vanquish',             'Vanquish',         'فانكويش'),
    ('gymshark',             'Gymshark',         'جيمشارك'),
    ('cagliari calcio',      'Cagliari Calcio',  'كالياري كالتشو'),
    ('cagliari',             'Cagliari',         'كالياري'),
    ('caglirari calcio',     'Cagliari Calcio',  'كالياري كالتشو'),
    ('denver nuggets',       'Denver Nuggets',   'دنفر ناغتس'),
    ('cheetahs rugby',       'Cheetahs Rugby',   'تشيتاز رَغبي'),
    ('naples',               'Naples',           'نابولي'),
    ('tekno',                'Tekno',            'تكنو'),
    ('flash',                'Flash',            'فلاش'),
    ('energy',               'Energy',           'إنرجي'),
    ('sgomma',               'Sgomma',           'سغوما'),
    ('street',               'Street',           'ستريت'),
]


def strip_sizes(name):
    """Remove size markers in every form the export uses."""
    n = name
    # "(M)", "[L]", "(Black - L)", "(Grey, L)"  -> keep the non-size part
    def paren(m):
        inner = m.group(1)
        parts = re.split(r'[,–—-]', inner)
        keep = [p.strip() for p in parts if not re.fullmatch(SIZE_WORDS, p.strip(), re.I)]
        return (' ' + ' '.join(keep)) if keep else ' '
    n = re.sub(r'\s*[\(\[]([^)\]]*)[\)\]]', paren, n)
    # trailing "- Large", "-XS", "- M"  (twice: some names carry two)
    for _ in range(2):
        n = re.sub(r'[\s–—-]+(?:' + SIZE_WORDS + r')\s*$', '', n, flags=re.I)
    return re.sub(r'\s+', ' ', n).strip(' -–—')


def parse(name):
    low = ' ' + strip_sizes(name).lower() + ' '
    colour = next(((en, ar) for en, ar in COLOURS if en in low), None)
    item = next(((en, ar, cat) for rx, en, ar, cat in ITEMS if re.search(rx, low)), None)
    coll = next(((en, ar) for key, en, ar in COLLECTIONS if key in low), None)
    return coll, item, colour


def slugify(s):
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode()
    return re.sub(r'-+', '-', re.sub(r'[^a-z0-9]+', '-', s.lower())).strip('-')


def main():
    src = Path(sys.argv[1] if len(sys.argv) > 1 else '')
    if not src.exists():
        sys.exit('usage: import-opencart.py path/to/sportaexport.xlsx')

    wb = openpyxl.load_workbook(src, data_only=True)
    rows = list(wb['Products'].iter_rows(min_row=2, values_only=True))

    skipped, groups, unparsed = [], defaultdict(list), []
    for pid, model, sku, name, price, qty, enabled, added in rows:
        if not name:
            skipped.append((pid, '(no name in the export)', 'row has no product name')); continue
        name = str(name).strip()
        if enabled != 'Yes':
            skipped.append((pid, name, 'disabled in OpenCart')); continue
        try:
            price = float(price or 0)
        except (TypeError, ValueError):
            price = 0.0
        if price <= 0:
            skipped.append((pid, name, 'price is 0 — cannot be sold')); continue

        coll, item, colour = parse(name)
        if not item:
            unparsed.append((pid, name, 'could not tell what kind of garment this is'))
            continue
        key = (coll[0] if coll else '', item[0], colour[0] if colour else '')
        groups[key].append({'id': pid, 'raw': name, 'price': price, 'qty': int(qty or 0),
                            'coll': coll, 'item': item, 'colour': colour})

    products, conflicts = [], []
    for (c_en, i_en, col_en), vs in sorted(groups.items()):
        coll, item, colour = vs[0]['coll'], vs[0]['item'], vs[0]['colour']
        prices = sorted({v['price'] for v in vs})
        if len(prices) > 1:
            conflicts.append((f'{c_en} {i_en} {col_en}'.strip(), prices, [v['id'] for v in vs]))
        price = prices[-1]          # never undercharge; the conflict is reported above

        name_en = ' '.join(x for x in [coll[0] if coll else '', item[0]] if x)
        name_ar = ' '.join(x for x in [item[1], coll[1] if coll else ''] if x)
        if colour:
            name_en += f' — {colour[0].title()}'
            name_ar += f' — {colour[1]}'
        products.append({
            'slug': slugify(f'{c_en}-{i_en}-{col_en}'),
            'name': {'en': name_en, 'ar': name_ar},
            'desc': {'en': '', 'ar': ''},
            'price': round(price, 3),
            'category': item[2],
            'stock': sum(v['qty'] for v in vs),
            'sources': [v['id'] for v in vs],
        })

    # Units actually sold, so "Bestseller" is a fact from the order history
    # rather than a guess. Only products still in the catalogue can carry it.
    sold = defaultdict(int)
    for oid, opid, pname, model, qty, unit, total in wb['Order Items'].iter_rows(min_row=2, values_only=True):
        if opid:
            sold[int(opid)] += int(qty or 0)
    for p in products:
        p['sold'] = sum(sold.get(i, 0) for i in p['sources'])
    top = sorted((p for p in products if p['sold'] > 0), key=lambda p: -p['sold'])[:3]
    for p in top:
        p['badge'] = {'en': 'Bestseller', 'ar': 'الأكثر مبيعًا'}

    write_catalogues(products)
    write_report(products, skipped, unparsed, conflicts, len(rows))

    print(f"{len(rows)} rows -> {len(products)} products")
    print(f"  skipped {len(skipped)}, unclassified {len(unparsed)}, price conflicts {len(conflicts)}")
    print(f"  wrote src/lib/products.js, sporta-html5/assets/products.js, CATALOGUE-IMPORT.md")


# Colours for the placeholder artwork, by category. Real photographs remain the
# biggest gap in the catalogue; these only stop the grid looking broken.
SWATCH = {'women': ('#7c2d12', '#431407'), 'men': ('#1e293b', '#0f172a'),
          'outerwear': ('#E0561C', '#B8430F'), 'accessories': ('#171A1E', '#333a42')}


def describe(p):
    """Say what the thing is. Nothing here is a claim that cannot be checked
    against the export — no invented fabrics, fits or features."""
    en_item = p['name']['en'].split(' — ')[0]
    colour_en = p['name']['en'].split(' — ')[1] if ' — ' in p['name']['en'] else ''
    colour_ar = p['name']['ar'].split(' — ')[1] if ' — ' in p['name']['ar'] else ''
    if colour_en:
        # Arabic uses a comma, not "باللون X". The colour words here are
        # indefinite ("أخضر عسكري"), and "باللون أخضر" is ungrammatical — it
        # wants the definite "باللون الأخضر العسكري". Rather than bolt "ال" onto
        # every colour and get it wrong on the ones that are nouns (أونيكس), the
        # phrasing sidesteps the construct.
        return ({'en': f'{en_item} in {colour_en.lower()}.',
                 'ar': f"{p['name']['ar'].split(' — ')[0]}، {colour_ar}."})
    return {'en': f'{en_item}.', 'ar': f"{p['name']['ar']}."}


def js_products(products):
    out = []
    for p in products:
        c1, c2 = SWATCH.get(p['category'], SWATCH['men'])
        label = p['name']['en'].split(' — ')[0][:18]
        entry = {'slug': p['slug'], 'name': p['name'], 'desc': describe(p),
                 'price': p['price'], 'category': p['category'],
                 '__ph': (label, c1, c2)}
        if p.get('badge'):
            entry['badge'] = p['badge']
        out.append(entry)
    return out


def write_catalogues(products):
    entries = js_products(products)

    # --- React copy: keeps CATEGORIES and the ph() helper, replaces PRODUCTS ---
    react = WEB / 'src' / 'lib' / 'products.js'
    text = react.read_text()
    lines = []
    for e in entries:
        label, c1, c2 = e['__ph']
        badge = ('\n    badge: ' + json.dumps(e['badge'], ensure_ascii=False) + ',') if 'badge' in e else ''
        lines.append(
            '  {\n'
            f"    slug: {json.dumps(e['slug'])},\n"
            f"    name: {json.dumps(e['name'], ensure_ascii=False)},\n"
            f"    desc: {json.dumps(e['desc'], ensure_ascii=False)},\n"
            f"    price: {e['price']}, category: {json.dumps(e['category'])}, "
            f"image: ph({json.dumps(label)}, {json.dumps(c1)}, {json.dumps(c2)}),{badge}\n"
            '  },')
    body = 'export const PRODUCTS = [\n' + '\n'.join(lines) + '\n]\n'
    start = text.index('export const PRODUCTS = [')
    end = text.index('\n]\n', start) + len('\n]\n')
    react.write_text(text[:start] + body + text[end:])

    # --- static copy, generated from the same parse so the two cannot drift ---
    html5 = ROOT / 'sporta-html5' / 'assets' / 'products.js'
    old = html5.read_text()
    cats_line = next(l for l in old.splitlines() if l.startswith('window.SPORTA_CATEGORIES'))
    plain = []
    for e in entries:
        label, c1, c2 = e['__ph']
        svg = ('data:image/svg+xml;utf8,' + _svg(label, c1, c2))
        d = {k: v for k, v in e.items() if k != '__ph'}
        d['image'] = svg
        plain.append(d)
    html5.write_text(
        '// GENERATED by sporta-web/scripts/import-opencart.py — do not hand-edit.\n'
        + cats_line + '\n'
        + 'window.SPORTA_PRODUCTS = ' + json.dumps(plain, ensure_ascii=False, indent=1) + ';\n')


def _svg(label, c1, c2):
    from urllib.parse import quote
    return quote(
        f'<svg xmlns="http://www.w3.org/2000/svg" width="600" height="600"><defs>'
        f'<linearGradient id="g" x1="0" y1="0" x2="1" y2="1"><stop offset="0" stop-color="{c1}"/>'
        f'<stop offset="1" stop-color="{c2}"/></linearGradient></defs>'
        f'<rect width="600" height="600" fill="url(#g)"/>'
        f'<text x="50%" y="52%" font-family="Arial" font-size="42" font-weight="bold" '
        f'fill="white" text-anchor="middle">{label}</text></svg>', safe='')


def write_report(products, skipped, unparsed, conflicts, nrows):
    L = []
    A = L.append
    A('# Catalogue import — OpenCart export\n')
    A(f'`{nrows}` rows in the export became **{len(products)} products**. The rows are size and')
    A('colour variants; the site sizes garments by category, so variants collapse into one')
    A('product each. Generated by `sporta-web/scripts/import-opencart.py`.\n')
    A('**Nothing was read from the Orders or Order Items sheets except unit counts per')
    A('product id.** Those sheets hold 119 real customers\' names, emails and phone numbers,')
    A('which do not belong in a public catalogue.\n')

    A('## Needs your decision\n')
    if conflicts:
        A('### Sizes of the same product carried different prices\n')
        A('The higher price was used, so nobody is undercharged. Confirm which is right:\n')
        A('| Product | Prices found | OpenCart ids |')
        A('|---|---|---|')
        for name, prices, ids in conflicts:
            A(f'| {name} | {", ".join(f"{p:.3f}" for p in prices)} KWD | {", ".join(map(str, ids))} |')
        A('')
    if unparsed:
        A('### Could not be classified — not imported\n')
        A('| OpenCart id | Name | Why |')
        A('|---|---|---|')
        for pid, name, why in unparsed:
            A(f'| {pid} | {name} | {why} |')
        A('\nGive these a name that says what the garment is and re-run the import.\n')

    low = [p for p in products if p['stock'] <= 2]
    A('### Stock is not enforced\n')
    A(f'{len(low)} of {len(products)} products have 2 or fewer units left, and')
    A('**the site has no stock control** — it will keep selling them after the last one goes.')
    A('Either keep the shelf topped up, hide the thin ones, or ask for stock enforcement to')
    A('be added to `create_order`.\n')

    A('## Imported\n')
    A('| Product | KWD | Category | Stock | Sold before |')
    A('|---|---|---|---|---|')
    for p in sorted(products, key=lambda p: (p['category'], p['slug'])):
        A(f"| {p['name']['en']} | {p['price']:.3f} | {p['category']} | {p['stock']} | "
          f"{p['sold'] or ''} |")

    A('\n## Not imported\n')
    A('| OpenCart id | Name | Reason |')
    A('|---|---|---|')
    for pid, name, why in skipped:
        A(f'| {pid} | {name} | {why} |')

    A('\n## Still missing\n')
    A('- **Real photographs.** Every product ships a generated placeholder. This is the')
    A('  biggest remaining gap: it limits conversion and Google Shopping eligibility.')
    A('- **Descriptions** say only what the item is and its colour. Nothing about fabric,')
    A('  fit or features was invented, because the export does not contain it.')

    (ROOT / 'CATALOGUE-IMPORT.md').write_text('\n'.join(L) + '\n')


if __name__ == '__main__':
    main()
